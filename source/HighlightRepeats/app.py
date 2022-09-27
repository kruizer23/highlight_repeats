#!/usr/bin/env python3
"""
Python program that highlights repeating text in a spreadsheet column.

Example usage:
   HighlightRepeats.py mysource.xlsx filetocreate.xlsx 2
"""
__docformat__ = 'reStructuredText'


# Imports
import argparse
import logging
import openpyxl
from openpyxl.styles import PatternFill


# Program return codes
RESULT_OK = 0
RESULT_ERROR_INPUT_FILE = 1
RESULT_ERROR_COLUMN = 2
RESULT_ERROR_ROWS = 3


def main():
    """
    Entry point into the program.
    """
    # Handle command line parameters.

    parser = argparse.ArgumentParser(description="Python program that highlights repeating text in a\r\n" +
                                                 "spreadsheet column.\r\n",
                                     formatter_class=argparse.RawDescriptionHelpFormatter)
    # Add mandatory command line parameters.
    parser.add_argument('infile', type=str, help='input spreadsheet file')
    parser.add_argument('outfile', type=str, help='output spreadsheet file')
    parser.add_argument('column', type=int, help='column to check')
    # Add optional command line parameters.
    parser.add_argument('-d', '--debug', action='store_true', dest='debug_enabled', default=False,
                        help='enable debug messages on the standard output')
    # Perform actual command line parameter parsing.
    args = parser.parse_args()

    # Enable debug logging level if requested.
    if args.debug_enabled:
        logging.basicConfig(level=logging.DEBUG)

    # Set the configuration values that were specified on the command line.
    cfg_in_file = args.infile;
    cfg_out_file = args.outfile;
    cfg_column = args.column;

    # Display configuration values for debugging purposes.
    logging.debug('Info: Input file: {}'.format(cfg_in_file))
    logging.debug('Info: Output file: {}'.format(cfg_out_file))
    logging.debug('Info: Column: {}'.format(cfg_column))

    # Open the workbook and get the active sheet.
    try:
        workbook = openpyxl.load_workbook(cfg_in_file)
    except Exception as e:
        logging.debug('Error: Cannot load workbook: {}'.format(e))
        return RESULT_ERROR_INPUT_FILE
    sheet = workbook.active

    # Make sure the selected column is a valid number.
    if cfg_column < 1:
        workbook.close()
        logging.debug('Error: Column ({}) cannot be smaller than one'.format(cfg_column))
        return RESULT_ERROR_COLUMN

    # Make sure the selected column is actually present
    if cfg_column > sheet.max_column:
        workbook.close()
        logging.debug('Error: Column ({}) larger than max columns ({})'.format(cfg_column, sheet.max_column))
        return RESULT_ERROR_COLUMN

    # Make sure the sheet has more then one row.
    if sheet.max_row <= 1:
        logging.debug('Error: Worksheet should have at least two rows.')
        return RESULT_ERROR_ROWS

    # Loop over all rows to look for the text in the configured column
    logging.debug('Info: Starting loop to look for repeats (iterations={}).'.format(sheet.max_row))
    for row_num in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_num, column=cfg_column)
        str_to_check = cell.value
        if str_to_check:
            # Check one specific string.
            check_column_for_repeats(sheet, cfg_column, str_to_check)

    # Save and the resulting workbook and close it.
    logging.debug('Info: Saving workbook to {}.'.format(cfg_out_file))
    workbook.save(cfg_out_file)
    logging.debug('Info: Closing workbook.')
    workbook.close()

    # Return exit code for success.
    return RESULT_OK


def check_column_for_repeats(sheet, column, str_to_check):
    """
    Loops over all cells of the configured column in the specified worksheet. If the cell
    contains the string to check, then it highlights the cell, except for the first entry.
    So it only highlights the repeats and not the first time the string to check is
    encountered.

    :param sheet: Worksheet object.
    :param column: Column to check.
    :param str_to_check: String to check.
    """
    # Initialize locals
    ok_to_highlight = False

    # Loop through all cells in the column to check
    for row_num in range(1, sheet.max_row + 1):
        cell = sheet.cell(row=row_num, column=column)
        # Does this call contain the string?
        if cell.value == str_to_check:
            # Okay to highlight this one?
            if ok_to_highlight:
                cell.fill = PatternFill("solid", start_color="FFD800")
            # Match found so update flag.
            ok_to_highlight = True


if __name__ == "__main__":
    main()
